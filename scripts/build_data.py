#!/usr/bin/env python3
"""Rebuild app data JSONs from official source publications in source-data/.
Usage: python3 scripts/build_data.py
Adjust FILE constants when new annual publications are released."""
import json, warnings, os
warnings.filterwarnings("ignore")
import openpyxl

SRC = os.path.join(os.path.dirname(__file__), "..", "source-data")
OUT = os.path.join(os.path.dirname(__file__), "..", "data")
os.makedirs(OUT, exist_ok=True)

def num(x):
    try: return float(str(x).replace(",", "."))
    except: return None

def build_uk():
    f = os.path.join(SRC, "nhpps-25-26-pay-award.xlsx")
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb["1 APC & OPROC"]
    out = {}
    for r in ws.iter_rows(min_row=7, values_only=True):
        code = str(r[1] or "")
        if not code.startswith(("HN", "HC")): continue
        out[code] = {"name": str(r[2] or "").strip(), "elective": num(r[4]), "non_elective": num(r[6])}
    return out

def build_us():
    base = 6752.61 + 524.15  # FY2026 operating + capital standardized amounts
    wb = openpyxl.load_workbook(os.path.join(SRC, "CMS-1833-F Table 5.xlsx"), data_only=True)
    ws = wb[wb.sheetnames[0]]
    out = {}
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r[0]: continue
        try:
            w = float(r[6]); alos = float(r[9])
        except (ValueError, TypeError, IndexError): continue
        out[str(r[0]).zfill(3)] = {"title": str(r[5]).strip(), "weight": w,
                                   "payment_usd": round(w * base), "alos": alos}
    return out

def build_de():
    wb = openpyxl.load_workbook(os.path.join(SRC, "g-drg-katalog-2025.xlsx"), read_only=True, data_only=True)
    ws = wb["Hauptabteilungen"]
    out = {}
    for r in ws.iter_rows(min_row=8, values_only=True):
        code = str(r[0] or "")
        if not code.startswith("I") or len(code) < 4: continue
        out[code] = {"desc": str(r[2] or "").strip(), "weight": num(r[3]), "alos": num(r[5])}
    return out

def build_uae():
    wb = openpyxl.load_workbook(os.path.join(SRC, "doh-drg-weights.xlsx"), read_only=True, data_only=True)
    ws = wb["WeightUpdate"]
    out = {}
    for r in ws.iter_rows(min_row=6, values_only=True):
        if not r[0]: continue
        try: w = float(r[9])
        except (ValueError, TypeError): continue
        out[str(r[0])] = {"desc": str(r[2] or ""), "severity": r[4], "weight": w}
    return out

def build_uk_activity():
    f = os.path.join(SRC, "NCC National Schedule_Supressed_2024_25", "Admitted Patient Care.xlsx")
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    from collections import defaultdict
    act = defaultdict(lambda: defaultdict(int))
    for r in ws.iter_rows(min_row=2, values_only=True):
        code = str(r[2] or "")
        if not code.startswith(("HN", "HC")): continue
        n = r[4]
        if isinstance(n, (int, float)): act[code][str(r[0])] += int(n)
    return {c: dict(v) for c, v in act.items()}

if __name__ == "__main__":
    for name, fn in [("uk_hrg_prices", build_uk), ("us_drg", build_us), ("de_gdrg", build_de),
                     ("uae_irdrg", build_uae), ("uk_activity", build_uk_activity)]:
        path = os.path.join(OUT, f"{name}.json")
        data = fn()
        json.dump(data, open(path, "w"), indent=1, ensure_ascii=False)
        print(f"{path}: {len(data)} rows")
